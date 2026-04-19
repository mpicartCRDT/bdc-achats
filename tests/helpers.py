from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook


def workbook_bytes(wb: Workbook) -> BytesIO:
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def make_parametrage_workbook(
    include_required: bool = True,
    include_supplier_delay_columns: bool = False,
) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Acheteurs")
    if include_required:
        ws.append(["code_acheteur", "nom_acheteur", "email_m365", "code_format_ph", "actif"])
        ws.append(["MATHIEU", "Mathieu Picart", "m.picart@example.test", "MATHIEU_V1", "OUI"])
    else:
        ws.append(["code_acheteur", "actif"])
        ws.append(["MATHIEU", "OUI"])

    ws = wb.create_sheet("Correspondance_Feuille_Usine")
    ws.append(["code_format_ph", "nom_feuille", "code_usine", "nom_usine", "actif"])
    ws.append(["MATHIEU_V1", "SPI 1", "KB", "Cabannes", "OUI"])

    ws = wb.create_sheet("Fournisseurs")
    supplier_columns = ["code_fournisseur", "nom_fournisseur", "actif"]
    if include_supplier_delay_columns:
        supplier_columns.extend(
            [
                "type_bdc",
                "code_delai_KB",
                "code_delai_C9",
                "code_delai_GNS",
                "code_delai_VNC",
                "mode_quantite_bdc",
            ]
        )
    ws.append(supplier_columns)
    suppliers = [
        ["AGRICOLA_PROGRES", "Agricola Progres", "OUI"],
        ["GROUPE_SAVEUR_CLOS", "Groupe Saveur des Clos", "OUI"],
        ["ROCA_DISTRIBUTION", "Roca Distribution", "OUI"],
    ]
    for supplier in suppliers:
        if include_supplier_delay_columns:
            supplier.extend(["Départ", "A-B", "A-C", "A-D", "A-B", "standard"])
        ws.append(supplier)

    ws = wb.create_sheet("Modeles_Mails")
    ws.append(["code_modele_mail", "code_fournisseur", "objet_modele", "corps_modele", "actif"])
    ws.append(
        [
            "MT_DEFAULT_FR",
            "*",
            "Programme semaine {{week_number}}",
            "Bonjour {{supplier_name}},\n\nCi-joint le programme de la semaine {{week_number}}.\n\nCordialement,\n{{buyer_name}}",
            "OUI",
        ]
    )

    ws = wb.create_sheet("Alias_Fournisseurs")
    ws.append(["alias", "code_fournisseur", "actif"])
    ws.append(["Ag Prog", "AGRICOLA_PROGRES", "OUI"])
    ws.append(["GSC", "GROUPE_SAVEUR_CLOS", "OUI"])
    ws.append(["Roca", "ROCA_DISTRIBUTION", "OUI"])

    ws = wb.create_sheet("Produits")
    ws.append(["code_produit", "nom_produit", "actif"])
    ws.append(["FRISEE", "Frisée", "OUI"])
    ws.append(["FDC_ROUGE", "FDC Rouge", "OUI"])
    ws.append(["TOMATE_PULPA", "T. Pulpa", "OUI"])

    ws = wb.create_sheet("Alias_Produits")
    ws.append(["alias", "code_produit", "actif"])
    ws.append(["Frisée", "FRISEE", "OUI"])
    ws.append(["F de Ch Rouge", "FDC_ROUGE", "OUI"])
    ws.append(["Tomate pulpa", "TOMATE_PULPA", "OUI"])

    return workbook_bytes(wb)


def make_bdc_template_workbook() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Feuil1"
    ws["C1"] = "Bon de commande"
    ws["G1"] = "Fournisseur"
    ws["C2"] = "Programme Semaine :"
    ws["F2"] = "XX"
    ws["C3"] = "Mail :"
    ws["F3"] = "XXXXXX@XXXXXXX.XX"
    ws["A4"] = "EN JOUR DE DEPART"
    ws["G4"] = "Transport"
    for block_index in range(8):
        start_row = 5 + block_index * 5
        ws.cell(start_row, 1).value = f"=A{start_row + 5}+7"
        ws.cell(start_row, 2).value = "Cabannes"
        ws.cell(start_row + 1, 2).value = "Chāteauneuf"
        ws.cell(start_row + 2, 2).value = "Genas"
        ws.cell(start_row + 3, 2).value = "VNC"
    return workbook_bytes(wb)
