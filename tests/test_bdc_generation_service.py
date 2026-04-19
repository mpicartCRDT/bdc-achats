from __future__ import annotations

from dataclasses import replace
from datetime import date
from io import BytesIO
import unittest

from openpyxl import load_workbook
from openpyxl.styles import Font

from parsers.mathieu_parser import MathieuParser
from repositories.parametrage_repository import ParametrageRepository
from services.bdc_generation_service import BDCGenerationService
from tests.helpers import (
    make_bdc_template_workbook,
    make_parametrage_workbook,
    workbook_bytes,
)
from tests.test_mathieu_parser import MathieuParserTest


class BDCGenerationServiceTest(unittest.TestCase):
    def setUp(self):
        self.references = ParametrageRepository().load(make_parametrage_workbook())
        self.references.suppliers_by_id["AGRICOLA_PROGRES"] = replace(
            self.references.suppliers_by_id["AGRICOLA_PROGRES"],
            bdc_type="Livraison",
            delay_codes_by_plant={"KB": "A-B"},
        )
        self.references.suppliers_by_id["GROUPE_SAVEUR_CLOS"] = replace(
            self.references.suppliers_by_id["GROUPE_SAVEUR_CLOS"],
            bdc_type="Départ",
            delay_codes_by_plant={"KB": "A-B"},
        )
        self.parser = MathieuParser(self.references)
        self.service = BDCGenerationService(self.references)

    def test_generates_one_workbook_per_recognized_supplier(self):
        ph_stream = MathieuParserTest()._make_ph_workbook()
        parsed = self.parser.parse(ph_stream, "PH test.xlsx")

        result = self.service.generate(make_bdc_template_workbook(), parsed.rows)

        self.assertEqual(
            sorted(file.supplier_id for file in result.files),
            ["AGRICOLA_PROGRES", "GROUPE_SAVEUR_CLOS"],
        )
        agricola = next(
            file for file in result.files if file.supplier_id == "AGRICOLA_PROGRES"
        )
        self.assertEqual(agricola.filename, "Agricola Progres S16.xlsx")
        wb = load_workbook(BytesIO(agricola.content), data_only=True)
        ws = wb.active

        self.assertEqual(ws["G1"].value, "Agricola Progres")
        self.assertEqual(ws["A4"].value, "EN JOUR DE LIVRAISON")
        self.assertEqual(ws["F2"].value, 16)
        self.assertEqual(ws["C4"].value, "Frisée")
        self.assertEqual(self._date_text(ws["A5"].value), "2026-04-11")
        self.assertEqual(self._date_text(ws["A10"].value), "2026-04-13")
        self.assertEqual(self._date_text(ws["A15"].value), "2026-04-14")
        self.assertEqual(self._date_text(ws["A20"].value), "2026-04-15")
        self.assertEqual(self._date_text(ws["A25"].value), "2026-04-16")
        self.assertEqual(self._date_text(ws["A30"].value), "2026-04-17")
        self.assertEqual(self._date_text(ws["A35"].value), "2026-04-18")
        self.assertEqual(self._date_text(ws["A40"].value), "2026-04-20")
        self.assertEqual(ws["C10"].value, 2)

    def test_depart_supplier_uses_delay_code_and_never_sunday(self):
        parsed = self.parser.parse(MathieuParserTest()._make_ph_workbook(), "PH test.xlsx")

        result = self.service.generate(make_bdc_template_workbook(), parsed.rows)

        gsc = next(
            file for file in result.files if file.supplier_id == "GROUPE_SAVEUR_CLOS"
        )
        wb = load_workbook(BytesIO(gsc.content), data_only=True)
        ws = wb.active

        self.assertEqual(ws["A4"].value, "EN JOUR DE DEPART")
        self.assertEqual(self._date_text(ws["A5"].value), "2026-04-11")
        self.assertEqual(ws["C5"].value, 1)
        self.assertIsNone(ws["C10"].value)

    def test_builds_zip_archive(self):
        parsed = self.parser.parse(MathieuParserTest()._make_ph_workbook(), "PH test.xlsx")
        result = self.service.generate(make_bdc_template_workbook(), parsed.rows)

        archive = self.service.build_zip(result.files)

        self.assertGreater(len(archive), 0)

    def test_fixed_bdc_dates_are_based_on_ph_delivery_week(self):
        dates = self.service._fixed_bdc_dates(date(2026, 4, 20))

        self.assertEqual(
            [item.isoformat() for item in dates],
            [
                "2026-04-18",
                "2026-04-20",
                "2026-04-21",
                "2026-04-22",
                "2026-04-23",
                "2026-04-24",
                "2026-04-25",
                "2026-04-27",
            ],
        )

    def test_supplier_quantity_mode_can_force_weight_in_bdc(self):
        self.references.suppliers_by_id["AGRICOLA_PROGRES"] = replace(
            self.references.suppliers_by_id["AGRICOLA_PROGRES"],
            bdc_type="Livraison",
            quantity_mode="poids_prioritaire",
        )
        parsed = self.parser.parse(
            MathieuParserTest()._make_weight_and_palette_workbook(), "PH poids.xlsx"
        )

        result = self.service.generate(make_bdc_template_workbook(), parsed.rows)

        agricola = next(
            file for file in result.files if file.supplier_id == "AGRICOLA_PROGRES"
        )
        wb = load_workbook(BytesIO(agricola.content), data_only=True)
        ws = wb.active
        self.assertEqual(ws["C10"].value, "120kg")

    def test_generated_values_are_forced_to_black_font(self):
        template = make_bdc_template_workbook()
        wb = load_workbook(template)
        ws = wb.active
        ws["C10"].font = Font(color="FFFF0000")
        template = workbook_bytes(wb)
        parsed = self.parser.parse(MathieuParserTest()._make_ph_workbook(), "PH test.xlsx")

        result = self.service.generate(template, parsed.rows)

        agricola = next(
            file for file in result.files if file.supplier_id == "AGRICOLA_PROGRES"
        )
        wb = load_workbook(BytesIO(agricola.content), data_only=True)
        ws = wb.active
        self.assertEqual(ws["C10"].font.color.rgb, "FF000000")

    def test_product_headers_keep_template_font_color(self):
        template = make_bdc_template_workbook()
        wb = load_workbook(template)
        ws = wb.active
        ws["C4"].font = Font(color="FFFFFFFF")
        template = workbook_bytes(wb)
        parsed = self.parser.parse(MathieuParserTest()._make_ph_workbook(), "PH test.xlsx")

        result = self.service.generate(template, parsed.rows)

        agricola = next(
            file for file in result.files if file.supplier_id == "AGRICOLA_PROGRES"
        )
        wb = load_workbook(BytesIO(agricola.content), data_only=True)
        ws = wb.active
        self.assertEqual(ws["C4"].value, "Frisée")
        self.assertEqual(ws["C4"].font.color.rgb, "FFFFFFFF")

    def _date_text(self, value):
        return value.date().isoformat() if hasattr(value, "date") else value.isoformat()


if __name__ == "__main__":
    unittest.main()
