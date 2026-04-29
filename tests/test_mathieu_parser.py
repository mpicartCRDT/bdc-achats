from __future__ import annotations

import unittest
from datetime import datetime

from openpyxl import Workbook

from parsers.mathieu_parser import MathieuParser
from repositories.parametrage_repository import ParametrageRepository
from tests.helpers import make_parametrage_workbook, workbook_bytes


class MathieuParserTest(unittest.TestCase):
    def setUp(self):
        self.references = ParametrageRepository().load(make_parametrage_workbook())
        self.parser = MathieuParser(self.references)

    def test_parses_rows_and_tracks_ignored_sheets(self):
        result = self.parser.parse(self._make_ph_workbook(), "PH test.xlsx")

        self.assertEqual(result.detected_sheets, ["SPI 1", "Autre"])
        self.assertEqual(result.processed_sheets, ["SPI 1"])
        self.assertEqual(result.ignored_sheets, ["Autre"])

        rows = result.rows
        self.assertEqual(len(rows), 5)
        self.assertEqual(rows[0].source_sheet, "SPI 1")
        self.assertEqual(rows[0].source_row_index, 6)
        self.assertEqual(rows[0].product_id, "FRISEE")
        self.assertEqual(rows[0].supplier_id, "AGRICOLA_PROGRES")
        self.assertEqual(rows[0].qty_value, 2)
        self.assertEqual(rows[0].qty_unit, "pal")
        self.assertEqual(rows[0].qty_nbre_raw, 2)
        self.assertIsNone(rows[0].qty_poids_raw)
        self.assertEqual(rows[0].source_week_number, 16)

        inherited = rows[1]
        self.assertEqual(inherited.source_row_index, 7)
        self.assertEqual(inherited.product_id, "FRISEE")
        self.assertEqual(inherited.supplier_id, "GROUPE_SAVEUR_CLOS")

        after_blank = rows[2]
        self.assertEqual(after_blank.source_row_index, 9)
        self.assertIsNone(after_blank.product_id)
        self.assertTrue(after_blank.needs_review)
        self.assertIn("Produit absent", after_blank.review_reason)

        after_total = rows[3]
        self.assertEqual(after_total.source_row_index, 11)
        self.assertIsNone(after_total.product_id)
        self.assertTrue(after_total.needs_review)

        after_section_break = rows[4]
        self.assertEqual(after_section_break.source_row_index, 13)
        self.assertIsNone(after_section_break.product_id)
        self.assertTrue(after_section_break.needs_review)

    def test_flags_text_quantity_for_review(self):
        result = self.parser.parse(self._make_text_quantity_workbook(), "PH text.xlsx")

        self.assertEqual(len(result.rows), 1)
        row = result.rows[0]
        self.assertTrue(row.needs_review)
        self.assertIn("Quantité non numérique", row.review_reason)
        self.assertEqual(row.info_text, "120 -cess. 60kg C9")

    def _make_ph_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "SPI 1"
        ws.append(["Prévision semaine :", None, None, None, 16])
        ws.append(["Jour de réception SPI 1 :"])
        ws.append([None, None, None, "Lundi", datetime(2026, 4, 13), None, None])
        ws.append([None, None, "PP", "Fournisseur", "Nbre", "poids", "trpt"])
        ws.append([None, None, None, None, "pal", None, None])
        ws.append(["Frisée", None, None, "Ag Prog", 2, None, "T1"])
        ws.append([None, None, None, "GSC", 1, None, None])
        ws.append([None, None, None, None, None, None, None])
        ws.append([None, None, None, "Roca", 3, None, None])
        ws.append(["TOTAL", None, None, "Ag Prog", 99, None, None])
        ws.append([None, None, None, "Roca", 4, None, None])
        ws.append(["FINE", None, None, None, None, None, None])
        ws.append([None, None, None, "Roca", 5, None, None])

        ignored = wb.create_sheet("Autre")
        ignored.append(["Pas Mathieu"])
        return workbook_bytes(wb)

    def _make_ph_workbook_with_saturday_s_minus_1(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "SPI 1"
        ws.append(["Prévision semaine :", None, None, None, 16])
        ws.append(["Jour de réception SPI 1 :"])
        ws.append(
            [
                None,
                None,
                None,
                "Samedi",
                datetime(2026, 4, 11),
                None,
                None,
                None,
                "Lundi",
                datetime(2026, 4, 13),
                None,
                None,
            ]
        )
        ws.append(
            [
                None,
                None,
                "PP",
                "Fournisseur",
                "Nbre",
                "poids",
                "trpt",
                "PP",
                "Fournisseur",
                "Nbre",
                "poids",
                "trpt",
            ]
        )
        ws.append(["Frisée", None, None, "Ag Prog", 4, None, None, None, "Ag Prog", 2, None, None])
        return workbook_bytes(wb)

    def _make_text_quantity_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "SPI 1"
        ws.append([None, None, None, "Lundi", datetime(2026, 4, 13), None, None])
        ws.append([None, None, "PP", "Fournisseur", "Nbre", "poids", "trpt"])
        ws.append(["Tomate pulpa", None, None, "Ag Prog", "120 -cess. 60kg C9", None, None])
        return workbook_bytes(wb)

    def _make_weight_and_palette_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "SPI 1"
        ws.append([None, None, None, "Lundi", datetime(2026, 4, 13), None, None])
        ws.append([None, None, "PP", "Fournisseur", "Nbre", "poids", "trpt"])
        ws.append(["Tomate pulpa", None, None, "Ag Prog", 2, 120, None])
        return workbook_bytes(wb)


if __name__ == "__main__":
    unittest.main()
