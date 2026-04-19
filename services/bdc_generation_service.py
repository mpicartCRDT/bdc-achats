from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any, BinaryIO
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from domain.models import (
    BDCGenerationResult,
    GeneratedBDCFile,
    NormalizedPHRow,
    ReferenceData,
    SupplierRef,
)
from services.normalization_service import normalize_key


@dataclass(frozen=True)
class TemplateLayout:
    worksheet_name: str
    product_header_row: int
    first_product_col: int
    transport_col: int
    plant_col: int
    day_blocks: list[list[int]]


class BDCGenerationError(ValueError):
    """User-facing BDC generation error."""


class BDCGenerationService:
    FIRST_PRODUCT_COL = 3
    PLANT_COL = 2
    PLANT_ROWS_PER_DAY = 4
    FIXED_WEEK_OFFSETS = [-2, 0, 1, 2, 3, 4, 5, 7]
    DELAY_DAYS = {
        "a-a": 0,
        "a-b": 1,
        "a-c": 2,
        "a-d": 3,
        "a-e": 4,
    }

    def __init__(self, references: ReferenceData):
        self.references = references

    def generate(
        self,
        template_source: BinaryIO,
        rows: list[NormalizedPHRow],
    ) -> BDCGenerationResult:
        eligible, skipped = self._split_rows(rows)
        grouped = self._group_by_supplier(eligible)
        result = BDCGenerationResult(skipped_rows=skipped)

        if not grouped:
            result.warnings.append("Aucune ligne reconnue disponible pour générer un BDC.")
            return result

        template_bytes = self._read_template_bytes(template_source)
        delivery_monday = self._delivery_week_monday(rows)
        for supplier_id, supplier_rows in grouped.items():
            supplier = self.references.suppliers_by_id[supplier_id]
            generated = self._generate_supplier_file(
                template_bytes=template_bytes,
                supplier_rows=supplier_rows,
                supplier=supplier,
                delivery_monday=delivery_monday,
                skipped_count=len([row for row in skipped if row.supplier_id == supplier_id]),
            )
            result.files.append(generated)

        return result

    @staticmethod
    def build_zip(files: list[GeneratedBDCFile]) -> bytes:
        stream = BytesIO()
        with ZipFile(stream, "w", ZIP_DEFLATED) as archive:
            for generated_file in files:
                archive.writestr(generated_file.filename, generated_file.content)
        return stream.getvalue()

    def _split_rows(
        self, rows: list[NormalizedPHRow]
    ) -> tuple[list[NormalizedPHRow], list[NormalizedPHRow]]:
        eligible: list[NormalizedPHRow] = []
        skipped: list[NormalizedPHRow] = []
        for row in rows:
            supplier = (
                self.references.suppliers_by_id.get(row.supplier_id)
                if row.supplier_id
                else None
            )
            can_generate = (
                supplier is not None
                and supplier.generate_bdc
                and not supplier.ignore_if_detected
                and not row.needs_review
                and row.product_id is not None
                and row.qty_value is not None
                and row.date_source is not None
            )
            if can_generate:
                eligible.append(row)
            else:
                skipped.append(row)
        return eligible, skipped

    def _group_by_supplier(
        self, rows: list[NormalizedPHRow]
    ) -> dict[str, list[NormalizedPHRow]]:
        grouped: dict[str, list[NormalizedPHRow]] = {}
        for row in rows:
            grouped.setdefault(str(row.supplier_id), []).append(row)
        return grouped

    def _generate_supplier_file(
        self,
        template_bytes: bytes,
        supplier_rows: list[NormalizedPHRow],
        supplier: SupplierRef,
        delivery_monday: date,
        skipped_count: int,
    ) -> GeneratedBDCFile:
        wb = load_workbook(BytesIO(template_bytes))
        ws = wb.active
        layout = self._detect_layout(ws)

        products = self._products_for_rows(supplier_rows)
        self._ensure_product_columns(ws, layout, len(products))
        layout = self._detect_layout(ws)
        product_cols = self._write_product_headers(ws, layout, products)

        dates = self._fixed_bdc_dates(delivery_monday)
        if len(layout.day_blocks) < len(dates):
            raise BDCGenerationError(
                "Le template BDC doit contenir 8 blocs jours : samedi S-1, lundi, mardi, mercredi, jeudi, vendredi, samedi, lundi S+1."
            )

        self._write_header(
            ws,
            supplier.supplier_name,
            supplier.email_to,
            delivery_monday,
            supplier.bdc_type,
        )
        date_blocks = self._assign_dates_to_blocks(ws, layout, dates)
        plant_rows = self._map_plant_rows(ws, date_blocks)
        injected = self._inject_rows(
            ws, supplier_rows, supplier, product_cols, plant_rows, layout
        )

        output = BytesIO()
        wb.save(output)
        return GeneratedBDCFile(
            supplier_id=supplier.supplier_id,
            supplier_name=supplier.supplier_name,
            filename=self._make_filename(supplier.supplier_name, delivery_monday),
            content=output.getvalue(),
            injected_rows=injected,
            skipped_rows=skipped_count + max(0, len(supplier_rows) - injected),
            week_number=delivery_monday.isocalendar().week,
        )

    def _detect_layout(self, ws: Worksheet) -> TemplateLayout:
        product_header_row = None
        transport_col = None
        for row in ws.iter_rows():
            for cell in row:
                if normalize_key(cell.value) == "transport":
                    product_header_row = cell.row
                    transport_col = cell.column
                    break
            if product_header_row and transport_col:
                break

        if product_header_row is None or transport_col is None:
            raise BDCGenerationError(
                "Le template BDC doit contenir une colonne 'Transport'."
            )

        day_blocks = self._detect_day_blocks(ws, product_header_row)
        if not day_blocks:
            raise BDCGenerationError(
                "Le template BDC doit contenir des lignes usines en colonne B."
            )

        return TemplateLayout(
            worksheet_name=ws.title,
            product_header_row=product_header_row,
            first_product_col=self.FIRST_PRODUCT_COL,
            transport_col=transport_col,
            plant_col=self.PLANT_COL,
            day_blocks=day_blocks,
        )

    def _detect_day_blocks(self, ws: Worksheet, header_row: int) -> list[list[int]]:
        runs: list[list[int]] = []
        current_run: list[int] = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            plant_value = ws.cell(row_idx, self.PLANT_COL).value
            if self._is_known_plant_label(plant_value):
                current_run.append(row_idx)
            elif current_run:
                runs.append(current_run)
                current_run = []
        if current_run:
            runs.append(current_run)

        blocks: list[list[int]] = []
        for run in runs:
            for start in range(0, len(run), self.PLANT_ROWS_PER_DAY):
                block = run[start : start + self.PLANT_ROWS_PER_DAY]
                if len(block) == self.PLANT_ROWS_PER_DAY:
                    blocks.append(block)
        return blocks

    def _is_known_plant_label(self, value: Any) -> bool:
        key = normalize_key(value)
        if not key:
            return False
        plant_keys = set()
        for plant in self.references.plants_by_sheet.values():
            plant_keys.update(self._plant_keys(plant.plant_code, plant.plant_name))
        plant_keys.update(
            {
                "cabannes",
                "chateauneuf",
                "genas",
                "vnc",
                "saint-julien",
                "saint julien",
                "perpi (gns)",
                "perpi (vnc)",
            }
        )
        return key in plant_keys or any(key.startswith(f"{item} ") for item in plant_keys)

    def _ensure_product_columns(
        self, ws: Worksheet, layout: TemplateLayout, product_count: int
    ) -> None:
        capacity = max(0, layout.transport_col - layout.first_product_col)
        if product_count <= capacity:
            return

        extra = product_count - capacity
        insert_at = layout.transport_col
        ws.insert_cols(insert_at, extra)
        source_col = max(layout.first_product_col, insert_at - 1)
        for col_offset in range(extra):
            target_col = insert_at + col_offset
            self._copy_column_style(ws, source_col, target_col)
            ws.column_dimensions[get_column_letter(target_col)].width = ws.column_dimensions[
                get_column_letter(source_col)
            ].width

    def _copy_column_style(self, ws: Worksheet, source_col: int, target_col: int) -> None:
        for row_idx in range(1, ws.max_row + 1):
            source = ws.cell(row_idx, source_col)
            target = ws.cell(row_idx, target_col)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy(source.alignment)
            if source.fill:
                target.fill = copy(source.fill)
            if source.border:
                target.border = copy(source.border)

    def _write_product_headers(
        self,
        ws: Worksheet,
        layout: TemplateLayout,
        products: list[tuple[str, str]],
    ) -> dict[str, int]:
        mapping: dict[str, int] = {}
        for offset, (product_id, product_name) in enumerate(products):
            col_idx = layout.first_product_col + offset
            cell = ws.cell(layout.product_header_row, col_idx)
            cell.value = product_name
            mapping[product_id] = col_idx
        return mapping

    def _products_for_rows(
        self, rows: list[NormalizedPHRow]
    ) -> list[tuple[str, str]]:
        product_ids = sorted({str(row.product_id) for row in rows if row.product_id})
        products = []
        for product_id in product_ids:
            product = self.references.products_by_id[product_id]
            products.append((product_id, product.product_name))
        return sorted(products, key=lambda item: normalize_key(item[1]))

    def _delivery_week_monday(self, rows: list[NormalizedPHRow]) -> date:
        dates = [self._as_date(row.date_source) for row in rows if row.date_source]
        if not dates:
            raise BDCGenerationError("Aucune date PH exploitable pour générer les BDC.")
        first_delivery_date = min(dates)
        return first_delivery_date - timedelta(days=first_delivery_date.weekday())

    def _fixed_bdc_dates(self, delivery_monday: date) -> list[date]:
        return [
            delivery_monday + timedelta(days=offset)
            for offset in self.FIXED_WEEK_OFFSETS
        ]

    def _write_header(
        self,
        ws: Worksheet,
        supplier_name: str,
        supplier_email: str | None,
        delivery_monday: date,
        bdc_type: str | None,
    ) -> None:
        week_number = delivery_monday.isocalendar().week
        for row in ws.iter_rows():
            for cell in row:
                key = normalize_key(cell.value)
                if key == "fournisseur":
                    cell.value = supplier_name
                    self._force_black_font(cell)
                elif key == "xx" and week_number is not None:
                    cell.value = week_number
                    self._force_black_font(cell)
                elif isinstance(cell.value, str) and "@" in cell.value and supplier_email:
                    cell.value = supplier_email
                    self._force_black_font(cell)
        ws["A4"].value = self._bdc_type_label(bdc_type)

    def _bdc_type_label(self, bdc_type: str | None) -> str:
        if normalize_key(bdc_type) == "livraison":
            return "EN JOUR DE LIVRAISON"
        return "EN JOUR DE DEPART"

    def _assign_dates_to_blocks(
        self, ws: Worksheet, layout: TemplateLayout, dates: list[date]
    ) -> dict[date, list[int]]:
        assignments: dict[date, list[int]] = {}
        for date_value, block_rows in zip(dates, layout.day_blocks):
            self._write_block_date(ws, block_rows, date_value)
            assignments[date_value] = block_rows
        return assignments

    def _write_block_date(
        self, ws: Worksheet, block_rows: list[int], date_value: date
    ) -> None:
        for row_idx in block_rows:
            cell = ws.cell(row_idx, 1)
            if not isinstance(cell, MergedCell):
                cell.value = None

        start_cell = self._top_left_cell(ws, block_rows[0], 1)
        start_cell.value = date_value
        if start_cell.number_format in {None, "General"}:
            start_cell.number_format = "[$-fr-FR]dddd d mmmm yyyy"
        self._force_black_font(start_cell)

    def _top_left_cell(self, ws: Worksheet, row: int, column: int):
        coordinate = ws.cell(row, column).coordinate
        for merged_range in ws.merged_cells.ranges:
            if coordinate in merged_range:
                return ws.cell(merged_range.min_row, merged_range.min_col)
        return ws.cell(row, column)

    def _map_plant_rows(
        self, ws: Worksheet, date_blocks: dict[date, list[int]]
    ) -> dict[tuple[date, str], int]:
        mapping: dict[tuple[date, str], int] = {}
        for date_value, block_rows in date_blocks.items():
            for row_idx in block_rows:
                plant_label = ws.cell(row_idx, self.PLANT_COL).value
                label_key = normalize_key(plant_label)
                for plant in self.references.plants_by_sheet.values():
                    if label_key in self._plant_keys(plant.plant_code, plant.plant_name):
                        mapping[(date_value, plant.plant_code)] = row_idx
        return mapping

    def _inject_rows(
        self,
        ws: Worksheet,
        rows: list[NormalizedPHRow],
        supplier: SupplierRef,
        product_cols: dict[str, int],
        plant_rows: dict[tuple[date, str], int],
        layout: TemplateLayout,
    ) -> int:
        injected = 0
        transport_values: dict[tuple[date, str], list[str]] = {}
        for row in rows:
            date_value = self._target_bdc_date(row, supplier)
            target_row = plant_rows.get((date_value, row.plant_code))
            target_col = product_cols.get(str(row.product_id))
            if target_row is None or target_col is None:
                continue

            cell = ws.cell(target_row, target_col)
            cell.value = self._sum_cell_value(cell.value, self._bdc_cell_value(row, supplier))
            self._force_black_font(cell)
            injected += 1

            if row.transport_value is not None:
                key = (date_value, row.plant_code)
                transport_values.setdefault(key, []).append(str(row.transport_value).strip())

        for (date_value, plant_code), values in transport_values.items():
            target_row = plant_rows.get((date_value, plant_code))
            if target_row is None:
                continue
            transport_cell = ws.cell(target_row, layout.transport_col)
            self._append_transport(transport_cell, values)

        return injected

    def _bdc_quantity_value(
        self, row: NormalizedPHRow, supplier: SupplierRef
    ) -> float | None:
        if normalize_key(supplier.quantity_mode) == "poids_prioritaire":
            weight = self._to_number(row.qty_poids_raw)
            if weight is not None:
                return weight
        return row.qty_value

    def _bdc_cell_value(self, row: NormalizedPHRow, supplier: SupplierRef) -> float | str | None:
        if normalize_key(supplier.quantity_mode) == "poids_prioritaire":
            weight = self._to_number(row.qty_poids_raw)
            if weight is not None:
                return f"{self._format_quantity(weight)}kg"
        return row.qty_value

    def _format_quantity(self, value: float) -> str:
        return str(int(value)) if float(value).is_integer() else str(value).replace(".", ",")

    def _target_bdc_date(self, row: NormalizedPHRow, supplier: SupplierRef) -> date:
        delivery_date = self._as_date(row.date_source)
        if normalize_key(supplier.bdc_type) != "depart":
            return delivery_date

        delay_code = self._delay_code_for_plant(supplier, row.plant_code)
        delay_days = self.DELAY_DAYS.get(normalize_key(delay_code), 0)
        target_date = delivery_date - timedelta(days=delay_days)
        if target_date.weekday() == 6:
            target_date -= timedelta(days=1)
        return target_date

    def _delay_code_for_plant(self, supplier: SupplierRef, plant_code: str) -> str | None:
        candidates = [plant_code]
        if plant_code == "GS":
            candidates.append("GNS")
        if plant_code == "GNS":
            candidates.append("GS")
        for candidate in candidates:
            if candidate in supplier.delay_codes_by_plant:
                return supplier.delay_codes_by_plant[candidate]
        return None

    def _sum_cell_value(self, existing: Any, value: float | str | None) -> float | int | str | Any:
        if value is None:
            return existing
        if isinstance(value, str):
            return self._append_text_value(existing, value)
        existing_number = self._to_number(existing)
        total = (existing_number or 0) + value
        return int(total) if float(total).is_integer() else total

    def _append_text_value(self, existing: Any, value: str) -> str:
        if existing is None:
            return value
        existing_text = str(existing).strip()
        if not existing_text:
            return value
        if value in existing_text.split(" + "):
            return existing_text
        return f"{existing_text} + {value}"

    def _append_transport(self, cell, values: list[str]) -> None:
        distinct = []
        for value in values:
            if value and value not in distinct:
                distinct.append(value)
        if not distinct:
            return

        existing = str(cell.value).strip() if cell.value is not None else ""
        additions = [value for value in distinct if value not in existing]
        if not additions:
            return
        cell.value = " / ".join([part for part in [existing, *additions] if part])
        self._force_black_font(cell)

    def _force_black_font(self, cell) -> None:
        font = copy(cell.font)
        font.color = "FF000000"
        cell.font = font

    def _plant_keys(self, plant_code: str, plant_name: str) -> set[str]:
        base = {normalize_key(plant_code), normalize_key(plant_name)}
        if plant_code == "KB":
            base.add("cabannes")
        if plant_code == "C9":
            base.add("chateauneuf")
        if plant_code in {"GS", "GNS"}:
            base.update({"genas", "perpi (gns)", "gns"})
        if plant_code == "VNC":
            base.update({"vnc", "saint-julien", "saint julien", "perpi (vnc)"})
        return base

    def _read_template_bytes(self, source: BinaryIO) -> bytes:
        if not hasattr(source, "read"):
            raise BDCGenerationError(
                "Le template BDC doit être chargé depuis l'interface."
            )
        if hasattr(source, "seek"):
            source.seek(0)
        content = source.read()
        if not content:
            raise BDCGenerationError("Le template BDC chargé est vide.")
        return content

    def _make_filename(self, supplier_name: str, delivery_monday: date) -> str:
        week = delivery_monday.isocalendar().week
        clean_supplier = " ".join(str(supplier_name).split())
        safe_supplier = "".join(
            char if char.isalnum() or char in {" ", "-", "_"} else " "
            for char in clean_supplier
        )
        safe_supplier = " ".join(safe_supplier.split())
        return f"{safe_supplier} S{week:02d}.xlsx"

    def _as_date(self, value: date | datetime | None) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        raise BDCGenerationError("Une ligne BDC n'a pas de date exploitable.")

    def _to_number(self, value: Any) -> float | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        try:
            return float(str(value).replace(",", "."))
        except ValueError:
            return None
